"""
Step 1 of the evaluation workflow: run a fixed set of test queries
through a live FreshRAG backend and dump the ranked results into a
spreadsheet with an empty column for you to fill in human relevance
judgments.

Usage:
    # Make sure the backend is running first: ./run_backend.sh
    python experiments/run_experiment.py --tag baseline
    python experiments/run_experiment.py --tag linear_decay   # after changing FRESHNESS_DECAY=linear in .env and restarting the backend
    python experiments/run_experiment.py --tag rule_based     # after setting QUERY_UNDERSTANDING_MODE=rule-based

Each run produces experiments/results_<tag>.xlsx. Run it once per
configuration you want to compare (baseline, each decay function,
rule-based vs LLM query understanding, etc) -- that's what turns into
your Comparative Analysis section.

Note on --delay: free-tier search (DuckDuckGo) and free-tier LLM APIs
(Groq's free tier especially) can rate-limit under a rapid burst of 18
back-to-back queries, even with the retry/backoff already built into
search_api.py and generator/llm.py. Increasing --delay spaces requests
out and reduces (though doesn't guarantee eliminating) rate-limit
failures. If you still see occasional 502s after this, that's normal
for a free-tier batch run -- just re-run the script again afterward;
it overwrites the same file, so failed queries simply get retried.
"""

import argparse
import json
import sys
import time
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="2E5395", end_color="2E5395", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11)

RESULT_COLUMNS = [
    "run_tag", "query_id", "query", "category", "domain", "time_sensitive",
    "rank", "source_title", "source_url", "published_date",
    "semantic_score", "freshness_score", "authority_score", "credibility_score",
    "final_score", "human_relevance",
]


def run_queries(backend_url: str, queries: list[dict], top_k: int, delay: float) -> list[dict]:
    rows = []
    for q in queries:
        print(f"  querying: {q['query']!r} ...", file=sys.stderr)
        try:
            resp = requests.post(
                f"{backend_url}/query",
                json={"query": q["query"], "top_k": top_k},
                timeout=120,
            )
            resp.raise_for_status()
            data = resp.json()
        except requests.RequestException as exc:
            print(f"    FAILED: {exc}", file=sys.stderr)
            time.sleep(delay)
            continue

        analysis = data.get("analysis", {})
        for rank, r in enumerate(data.get("results", []), start=1):
            scores = r.get("scores", {})
            rows.append({
                "query_id": q["id"],
                "query": q["query"],
                "category": q.get("category", ""),
                "domain": analysis.get("domain", ""),
                "time_sensitive": analysis.get("time_sensitive", ""),
                "rank": rank,
                "source_title": r.get("source_title", ""),
                "source_url": r.get("source_url", ""),
                "published_date": r.get("published_date", ""),
                "semantic_score": scores.get("semantic", ""),
                "freshness_score": scores.get("freshness", ""),
                "authority_score": scores.get("authority", ""),
                "credibility_score": scores.get("credibility", ""),
                "final_score": scores.get("final", ""),
            })
        time.sleep(delay)  # be polite to free-tier search/LLM rate limits between queries
    return rows


def write_workbook(rows: list[dict], run_tag: str, out_path: Path) -> None:
    wb = Workbook()

    instr = wb.active
    instr.title = "Instructions"
    instr["A1"] = "How to fill in this spreadsheet"
    instr["A1"].font = Font(name="Calibri", bold=True, size=14)
    lines = [
        "",
        "Go to the 'Results' sheet. Each row is one ranked source for one query.",
        "Fill in the 'human_relevance' column (highlighted yellow) for every row using this scale:",
        "",
        "  3 = Excellent: directly answers the query, current and reliable",
        "  2 = Relevant: on-topic but not ideal (a bit outdated, or only partially answers it)",
        "  1 = Barely related: mentions the topic but doesn't really help",
        "  0 = Irrelevant: off-topic or wrong",
        "",
        "Once every row has a number in that column, save the file and run:",
        f"  python experiments/compute_metrics.py experiments/{out_path.name}",
        "",
        "Run this same process once per configuration you want to compare",
        "(e.g. different FRESHNESS_DECAY settings, or QUERY_UNDERSTANDING_MODE=rule-based",
        "vs auto), using a different --tag each time, then pass multiple result files to",
        "compute_metrics.py together to get a side-by-side comparison table.",
        "",
        "If some queries FAILED (502/500 errors) due to free-tier search or LLM",
        "rate limits, just re-run this same command again -- it overwrites the",
        "same file, so previously-successful rows get regenerated and failed",
        "ones get another chance.",
    ]
    for i, line in enumerate(lines, start=2):
        instr[f"A{i}"] = line
        instr[f"A{i}"].font = BODY_FONT
    instr.column_dimensions["A"].width = 100

    ws = wb.create_sheet("Results")
    for col_idx, col_name in enumerate(RESULT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        row["run_tag"] = run_tag
        for col_idx, col_name in enumerate(RESULT_COLUMNS, start=1):
            value = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            if col_name == "human_relevance":
                cell.fill = INPUT_FILL

    for col_idx, col_name in enumerate(RESULT_COLUMNS, start=1):
        width = 14 if col_name not in ("query", "source_title", "source_url") else 40
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--tag", required=True, help="Label for this run, e.g. 'baseline', 'linear_decay', 'rule_based'")
    parser.add_argument("--backend-url", default="http://localhost:8000")
    parser.add_argument("--top-k", type=int, default=8)
    parser.add_argument("--queries-file", default=str(Path(__file__).parent / "queries.json"))
    parser.add_argument("--delay", type=float, default=2.0,
                         help="Seconds to wait between queries (default 2.0). Increase this if you see "
                              "502 errors from search/LLM rate-limiting under free-tier keys.")
    args = parser.parse_args()

    queries = json.loads(Path(args.queries_file).read_text())["queries"]
    print(f"Running {len(queries)} queries against {args.backend_url} (tag={args.tag}, delay={args.delay}s) ...", file=sys.stderr)
    rows = run_queries(args.backend_url, queries, args.top_k, args.delay)

    if not rows:
        print("No results were returned for any query -- is the backend running and reachable?", file=sys.stderr)
        sys.exit(1)

    out_path = Path(__file__).parent / f"results_{args.tag}.xlsx"
    write_workbook(rows, args.tag, out_path)
    print(f"\nWrote {len(rows)} result rows to {out_path}", file=sys.stderr)
    print("Next: open it, fill in the 'human_relevance' column, save, then run compute_metrics.py on it.", file=sys.stderr)


if __name__ == "__main__":
    main()
