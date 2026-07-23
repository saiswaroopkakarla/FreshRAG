"""
Step 2 of the evaluation workflow: read one or more labeled result
spreadsheets (produced by run_experiment.py and filled in by hand) and
compute Precision@k, Recall@k (pooled), nDCG@k, and freshness-
satisfaction, both per-query and averaged, using the same
evaluation/metrics.py functions used elsewhere in the project.

Usage:
    python experiments/compute_metrics.py experiments/results_baseline.xlsx
    python experiments/compute_metrics.py experiments/results_baseline.xlsx experiments/results_linear_decay.xlsx ...

Output: experiments/metrics_summary.xlsx, with THREE summary views:
  - "Comparison Summary": each run averaged over whatever queries it
    actually has data for. If runs have different numbers of queries
    (e.g. one had a search failure another didn't), this is NOT a
    perfectly apples-to-apples comparison -- differences could partly
    reflect which queries got averaged, not the config change itself.
  - "Common-Queries Summary": each run averaged over ONLY the query_ids
    present in every run being compared. Fewer queries, but a fair,
    matched comparison -- this is the one to trust/report when run
    completeness differs across configs.
  - "Per-Query Metrics" / "Raw Judged Data": unchanged detail sheets.
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, str(Path(__file__).parent.parent))
from evaluation.metrics import ndcg_at_k, precision_at_k, recall_at_k, freshness_satisfaction  # noqa: E402

HEADER_FILL = PatternFill(start_color="2E5395", end_color="2E5395", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11)
RELEVANT_THRESHOLD = 2
K_VALUES = [3, 5, 8]


def load_run(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Results")
    missing = df["human_relevance"].isna()
    if missing.any():
        n = missing.sum()
        print(f"  WARNING: {n} row(s) in {path.name} have no human_relevance filled in -- "
              f"they will be excluded from metrics for their query.", file=sys.stderr)
    df = df.dropna(subset=["human_relevance"])
    df["human_relevance"] = df["human_relevance"].astype(int)
    df["rank"] = df["rank"].astype(int)
    return df


def compute_per_query_metrics(df: pd.DataFrame) -> pd.DataFrame:
    pooled_relevant_by_query = {
        query_id: set(group.loc[group["human_relevance"] >= RELEVANT_THRESHOLD, "source_url"])
        for query_id, group in df.groupby("query_id")
    }

    records = []
    for (run_tag, query_id), group in df.groupby(["run_tag", "query_id"]):
        group = group.sort_values("rank")
        retrieved_ids = group["source_url"].tolist()
        relevance_scores = group["human_relevance"].tolist()
        relevant_ids = set(group.loc[group["human_relevance"] >= RELEVANT_THRESHOLD, "source_url"])
        pooled_relevant_ids = pooled_relevant_by_query[query_id]

        row = {
            "run_tag": run_tag,
            "query_id": query_id,
            "query": group["query"].iloc[0],
            "category": group["category"].iloc[0],
            "n_judged": len(group),
            "n_relevant": len(relevant_ids),
        }
        for k in K_VALUES:
            if k <= len(retrieved_ids):
                row[f"precision@{k}"] = round(precision_at_k(retrieved_ids, relevant_ids, k), 3)
                row[f"recall@{k}"] = round(recall_at_k(retrieved_ids, pooled_relevant_ids, k), 3)
                row[f"ndcg@{k}"] = round(ndcg_at_k(relevance_scores, k), 3)
            else:
                row[f"precision@{k}"] = None
                row[f"recall@{k}"] = None
                row[f"ndcg@{k}"] = None
        row["freshness_satisfaction"] = round(
            freshness_satisfaction(group["freshness_score"].astype(float).tolist()), 3
        )
        row["avg_final_score"] = round(group["final_score"].astype(float).mean(), 3)
        records.append(row)
    return pd.DataFrame(records)


def compute_run_summary(per_query: pd.DataFrame, label: str = "run_tag") -> pd.DataFrame:
    metric_cols = [c for c in per_query.columns if c.startswith(("precision@", "recall@", "ndcg@"))] + [
        "freshness_satisfaction", "avg_final_score"
    ]
    summary = per_query.groupby(label)[metric_cols].mean().round(3).reset_index()
    n_queries = per_query.groupby(label)["query_id"].nunique().rename("n_queries")
    summary = summary.merge(n_queries, on=label)
    return summary


def compute_common_queries_summary(per_query: pd.DataFrame) -> tuple[pd.DataFrame, set]:
    """Restrict to only the query_ids present for EVERY run_tag, so the
    comparison is fair even when some runs are missing queries due to
    search/API failures during collection."""
    query_sets_per_run = per_query.groupby("run_tag")["query_id"].apply(set)
    common_queries = set.intersection(*query_sets_per_run.tolist()) if len(query_sets_per_run) > 0 else set()
    filtered = per_query[per_query["query_id"].isin(common_queries)]
    if filtered.empty:
        return pd.DataFrame(), common_queries
    return compute_run_summary(filtered), common_queries


def write_workbook(all_raw: pd.DataFrame, per_query: pd.DataFrame, summary: pd.DataFrame,
                    common_summary: pd.DataFrame, common_queries: set, out_path: Path) -> None:
    wb = Workbook()

    def write_df(ws, df: pd.DataFrame, start_row: int = 1):
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == start_row:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                else:
                    cell.font = BODY_FONT
        for c_idx, col_name in enumerate(df.columns, start=1):
            width = 42 if col_name in ("query",) else 16
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    ws_summary = wb.active
    ws_summary.title = "Comparison Summary"
    ws_summary["A1"] = "Full averages -- each run averaged over whatever queries IT has (n_queries may differ per run; see note below)"
    ws_summary["A1"].font = Font(name="Calibri", bold=True, size=12)
    write_df(ws_summary, summary, start_row=3)
    ws_summary.freeze_panes = "A4"
    for c_idx in range(1, len(summary.columns) + 1):
        ws_summary.column_dimensions[get_column_letter(c_idx)].width = 16

    ws_common = wb.create_sheet("Common-Queries Summary")
    ws_common["A1"] = f"Fair comparison -- restricted to the {len(common_queries)} query_id(s) present in EVERY run below. Use this one for reporting when n_queries differs above."
    ws_common["A1"].font = Font(name="Calibri", bold=True, size=12)
    if not common_summary.empty:
        write_df(ws_common, common_summary, start_row=3)
        ws_common.freeze_panes = "A4"
        for c_idx in range(1, len(common_summary.columns) + 1):
            ws_common.column_dimensions[get_column_letter(c_idx)].width = 16
    else:
        ws_common["A3"] = "No common queries across all loaded runs -- cannot compute a fair comparison."
        ws_common["A3"].font = BODY_FONT

    ws_perquery = wb.create_sheet("Per-Query Metrics")
    write_df(ws_perquery, per_query)
    ws_perquery.freeze_panes = "A2"

    ws_raw = wb.create_sheet("Raw Judged Data")
    write_df(ws_raw, all_raw)
    ws_raw.freeze_panes = "A2"

    wb.save(out_path)


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    paths = [Path(p) for p in sys.argv[1:]]
    frames = []
    for p in paths:
        if not p.exists():
            print(f"File not found: {p}", file=sys.stderr)
            sys.exit(1)
        print(f"Loading {p} ...", file=sys.stderr)
        frames.append(load_run(p))

    all_raw = pd.concat(frames, ignore_index=True)
    if all_raw.empty:
        print("No judged rows found across the given file(s) -- fill in human_relevance first.", file=sys.stderr)
        sys.exit(1)

    per_query = compute_per_query_metrics(all_raw)
    summary = compute_run_summary(per_query)
    common_summary, common_queries = compute_common_queries_summary(per_query)

    out_path = Path(__file__).parent / "metrics_summary.xlsx"
    write_workbook(all_raw, per_query, summary, common_summary, common_queries, out_path)

    print(f"\nWrote comparison summary to {out_path}", file=sys.stderr)
    print("\n=== Full averages (n_queries may differ per run) ===", file=sys.stderr)
    print(summary.to_string(index=False), file=sys.stderr)
    print(f"\n=== Common-queries-only averages ({len(common_queries)} shared queries) ===", file=sys.stderr)
    if not common_summary.empty:
        print(common_summary.to_string(index=False), file=sys.stderr)
    else:
        print("(no queries in common across all loaded runs)", file=sys.stderr)


if __name__ == "__main__":
    main()
